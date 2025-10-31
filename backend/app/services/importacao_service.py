# ============================================================================
# SERVICE - PROCESSAMENTO DE IMPORTAÇÃO DE INSUMOS
# ============================================================================
# Descrição: Serviço para processar arquivos Excel e importar insumos
# Data: 30/10/2025
# Autor: Will - Empresa: IOGAR
# ============================================================================

import openpyxl
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
from datetime import datetime
from sqlalchemy.orm import Session
from sqlalchemy import func
import json
import logging

from app.models.importacao_insumo import ImportacaoInsumo, StatusImportacao
from app.models.insumo import Insumo
from app.schemas.importacao_insumo import (
    PreviewImportacao,
    LogProcessamento,
    ItemLog
)

# Configurar logging
logger = logging.getLogger(__name__)


# ============================================================================
# CONSTANTES DE MAPEAMENTO
# ============================================================================

MAPEAMENTO_COLUNAS = {
    'CodigoProduto': 'codigo',
    'NomeProduto': 'nome',
    'PrecoCompra': 'preco_compra_real',
    'Unidade': 'unidade'
}

CONVERSAO_UNIDADES = {
    'LT': 'L',
    'UN': 'unidade',
    'KG': 'kg',
    'G': 'g',
    'ML': 'ml',
    'L': 'L'
}


# ============================================================================
# FUNÇÕES AUXILIARES
# ============================================================================

def converter_unidade(unidade_excel: str) -> str:
    """
    Converte unidade do Excel para o padrão do sistema.
    
    Args:
        unidade_excel: Unidade vinda do Excel (LT, UN, KG, etc.)
        
    Returns:
        str: Unidade no padrão do sistema
        
    Exemplos:
        LT -> L
        UN -> unidade
        KG -> kg
    """
    if not unidade_excel:
        return 'unidade'
    
    unidade_upper = str(unidade_excel).strip().upper()
    return CONVERSAO_UNIDADES.get(unidade_upper, unidade_excel.lower())


def encontrar_linha_cabecalho(worksheet) -> Optional[int]:
    """
    Encontra a linha do cabeçalho no Excel.
    Procura por palavras-chave como 'CodigoProduto', 'NomeProduto', etc.
    
    Args:
        worksheet: Planilha do openpyxl
        
    Returns:
        int: Número da linha do cabeçalho (1-indexed) ou None se não encontrar
    """
    for row_num, row in enumerate(worksheet.iter_rows(min_row=1, max_row=20), start=1):
        valores = [str(cell.value).strip() if cell.value else '' for cell in row]
        
        # Verificar se contém colunas esperadas
        if any('CodigoProduto' in v or 'NomeProduto' in v for v in valores):
            return row_num
    
    return None


def extrair_dados_linha(row, colunas_mapeadas: Dict[str, int]) -> Dict[str, Any]:
    """
    Extrai dados de uma linha do Excel usando o mapeamento de colunas.
    
    Args:
        row: Linha do Excel (tupla de células)
        colunas_mapeadas: Dicionário {nome_campo: índice_coluna}
        
    Returns:
        Dict com os dados extraídos e convertidos
    """
    dados = {}
    
    # Extrair código
    if 'codigo' in colunas_mapeadas:
        codigo_cell = row[colunas_mapeadas['codigo']]
        dados['codigo'] = str(codigo_cell.value).strip() if codigo_cell.value else None
    
    # Extrair nome
    if 'nome' in colunas_mapeadas:
        nome_cell = row[colunas_mapeadas['nome']]
        dados['nome'] = str(nome_cell.value).strip() if nome_cell.value else None
    
    # Extrair preço
    if 'preco_compra_real' in colunas_mapeadas:
        preco_cell = row[colunas_mapeadas['preco_compra_real']]
        try:
            preco = float(preco_cell.value) if preco_cell.value else 0.0
            dados['preco_compra_real'] = preco
        except (ValueError, TypeError):
            dados['preco_compra_real'] = 0.0
    
    # Extrair unidade
    if 'unidade' in colunas_mapeadas:
        unidade_cell = row[colunas_mapeadas['unidade']]
        unidade_raw = str(unidade_cell.value).strip() if unidade_cell.value else 'unidade'
        dados['unidade'] = converter_unidade(unidade_raw)
    
    return dados


def validar_dados_linha(dados: Dict[str, Any], linha_num: int) -> Tuple[bool, Optional[str]]:
    """
    Valida se os dados extraídos estão corretos.
    
    Args:
        dados: Dicionário com dados extraídos
        linha_num: Número da linha (para mensagem de erro)
        
    Returns:
        Tuple[bool, Optional[str]]: (válido, mensagem_erro)
    """
    # Validar código
    if not dados.get('codigo'):
        return False, f"Linha {linha_num}: Código do produto está vazio"
    
    # Validar nome
    if not dados.get('nome'):
        return False, f"Linha {linha_num}: Nome do produto está vazio"
    
    # Validar unidade
    if not dados.get('unidade'):
        return False, f"Linha {linha_num}: Unidade está vazia"
    
    return True, None


# ============================================================================
# CLASSE PRINCIPAL DO SERVICE
# ============================================================================

class ImportacaoService:
    """
    Service para processamento de importação de insumos via Excel.
    """
    
    def __init__(self, db: Session):
        """
        Inicializa o service com a sessão do banco de dados.
        
        Args:
            db: Sessão do SQLAlchemy
        """
        self.db = db
    
    # ========================================================================
    # MÉTODO: GERAR PREVIEW
    # ========================================================================
    
    def gerar_preview(
        self,
        caminho_arquivo: str,
        nome_arquivo: str
    ) -> PreviewImportacao:
        """
        Gera preview dos dados que serão importados.
        
        Args:
            caminho_arquivo: Caminho do arquivo Excel no servidor
            nome_arquivo: Nome original do arquivo
            
        Returns:
            PreviewImportacao: Schema com preview dos dados
            
        Raises:
            ValueError: Se arquivo inválido ou não puder ser lido
        """
        try:
            # Abrir arquivo Excel
            wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
            ws = wb.active
            
            # Encontrar linha do cabeçalho
            linha_cabecalho = encontrar_linha_cabecalho(ws)
            if not linha_cabecalho:
                raise ValueError("Não foi possível encontrar o cabeçalho no arquivo Excel")
            
            # Extrair colunas do cabeçalho
            header_row = list(ws.iter_rows(min_row=linha_cabecalho, max_row=linha_cabecalho))[0]
            colunas_detectadas = [
                str(cell.value).strip() if cell.value else f"Coluna_{i+1}"
                for i, cell in enumerate(header_row)
            ]
            
            # Mapear colunas
            colunas_mapeadas = {}
            for i, col in enumerate(colunas_detectadas):
                campo_sistema = MAPEAMENTO_COLUNAS.get(col)
                if campo_sistema:
                    colunas_mapeadas[campo_sistema] = i
            
            # Extrair primeiras 5 linhas de dados
            primeiras_linhas = []
            linha_dados_inicial = linha_cabecalho + 1
            
            for row in ws.iter_rows(min_row=linha_dados_inicial, max_row=linha_dados_inicial + 4):
                dados = extrair_dados_linha(row, colunas_mapeadas)
                if dados.get('codigo') and dados.get('nome'):
                    primeiras_linhas.append(dados)
            
            # Contar total de linhas
            total_linhas = ws.max_row - linha_cabecalho
            
            # Gerar avisos
            avisos = []
            if 'codigo' not in colunas_mapeadas:
                avisos.append("⚠️ Coluna 'CodigoProduto' não encontrada")
            if 'nome' not in colunas_mapeadas:
                avisos.append("⚠️ Coluna 'NomeProduto' não encontrada")
            if 'preco_compra_real' not in colunas_mapeadas:
                avisos.append("⚠️ Coluna 'PrecoCompra' não encontrada")
            if 'unidade' not in colunas_mapeadas:
                avisos.append("⚠️ Coluna 'Unidade' não encontrada")
            
            wb.close()
            
            return PreviewImportacao(
                nome_arquivo=nome_arquivo,
                total_linhas=total_linhas,
                colunas_detectadas=colunas_detectadas,
                primeiras_linhas=primeiras_linhas,
                mapeamento_colunas={
                    k: v for k, v in MAPEAMENTO_COLUNAS.items()
                },
                avisos=avisos
            )
            
        except Exception as e:
            logger.error(f"Erro ao gerar preview: {e}")
            raise ValueError(f"Erro ao processar arquivo: {str(e)}")
    
    # ========================================================================
    # MÉTODO: PROCESSAR IMPORTAÇÃO
    # ========================================================================
    
    def processar_importacao(
        self,
        importacao_id: int,
        restaurante_id: int
    ) -> Tuple[bool, str]:
        """
        Processa a importação e cria os insumos no banco de dados.
        
        Args:
            importacao_id: ID da importação a processar
            restaurante_id: ID do restaurante
            
        Returns:
            Tuple[bool, str]: (sucesso, mensagem)
        """
        # Buscar importação
        importacao = self.db.query(ImportacaoInsumo).filter(
            ImportacaoInsumo.id == importacao_id
        ).first()
        
        if not importacao:
            return False, "Importação não encontrada"
        
        # Atualizar status para processando
        importacao.status = StatusImportacao.PROCESSANDO
        importacao.data_inicio_processamento = datetime.now()
        self.db.commit()
        
        # Inicializar log
        log = LogProcessamento()
        
        try:
            # Abrir arquivo Excel
            wb = openpyxl.load_workbook(importacao.caminho_arquivo, data_only=True)
            ws = wb.active
            
            # Encontrar cabeçalho
            linha_cabecalho = encontrar_linha_cabecalho(ws)
            if not linha_cabecalho:
                raise ValueError("Cabeçalho não encontrado")
            
            # Mapear colunas
            header_row = list(ws.iter_rows(min_row=linha_cabecalho, max_row=linha_cabecalho))[0]
            colunas_mapeadas = {}
            for i, cell in enumerate(header_row):
                col_name = str(cell.value).strip() if cell.value else ''
                campo_sistema = MAPEAMENTO_COLUNAS.get(col_name)
                if campo_sistema:
                    colunas_mapeadas[campo_sistema] = i
            
            # Processar linhas
            linha_dados_inicial = linha_cabecalho + 1
            total_linhas = ws.max_row - linha_cabecalho
            
            for row_num, row in enumerate(
                ws.iter_rows(min_row=linha_dados_inicial),
                start=linha_dados_inicial
            ):
                # Extrair dados
                dados = extrair_dados_linha(row, colunas_mapeadas)
                
                # ========================================================================
                # FILTRO: Apenas códigos entre 5000 e 5999
                # ========================================================================
                try:
                    codigo_numero = int(dados.get('codigo', 0))
                    if codigo_numero < 5000 or codigo_numero > 5999:
                        log.ignorados.append(ItemLog(
                            linha=row_num,
                            tipo="ignorado",
                            mensagem=f"Código {dados['codigo']} fora da faixa permitida (5000-5999)",
                            dados=dados
                        ))
                        continue
                except (ValueError, TypeError):
                    log.erros.append(ItemLog(
                        linha=row_num,
                        tipo="erro",
                        mensagem=f"Código inválido: {dados.get('codigo')}",
                        dados=dados
                    ))
                    continue
                
                # Validar dados
                valido, erro = validar_dados_linha(dados, row_num)
                
                if not valido:
                    log.erros.append(ItemLog(
                        linha=row_num,
                        tipo="erro",
                        mensagem=erro,
                        dados=dados
                    ))
                    continue
                
                # Verificar se insumo já existe
                insumo_existente = self.db.query(Insumo).filter(
                    Insumo.restaurante_id == restaurante_id,
                    Insumo.codigo == dados['codigo']
                ).first()
                
                if insumo_existente:
                    log.ignorados.append(ItemLog(
                        linha=row_num,
                        tipo="ignorado",
                        mensagem=f"Insumo com código {dados['codigo']} já existe",
                        dados=dados
                    ))
                    continue
                
                # Criar insumo
                try:
                    novo_insumo = Insumo(
                        restaurante_id=restaurante_id,
                        importacao_id=importacao_id,
                        codigo=dados['codigo'],
                        nome=dados['nome'],
                        quantidade=1,  # Padrão
                        unidade=dados['unidade'],
                        preco_compra=int(dados['preco_compra_real'] * 100),  # Converter para centavos
                        grupo='',  # Padrão vazio
                        subgrupo='',  # Padrão vazio
                        eh_fornecedor_anonimo=True,
                        aguardando_classificacao=False
                    )
                    
                    self.db.add(novo_insumo)
                    
                    log.sucessos.append(ItemLog(
                        linha=row_num,
                        tipo="sucesso",
                        mensagem=f"Insumo '{dados['nome']}' importado com sucesso",
                        dados=dados
                    ))
                    
                except Exception as e:
                    log.erros.append(ItemLog(
                        linha=row_num,
                        tipo="erro",
                        mensagem=f"Erro ao criar insumo: {str(e)}",
                        dados=dados
                    ))
            
            wb.close()
            
            # Atualizar estatísticas da importação
            importacao.total_linhas = total_linhas
            importacao.linhas_processadas = len(log.sucessos)
            importacao.linhas_com_erro = len(log.erros)
            importacao.linhas_ignoradas = len(log.ignorados)
            importacao.log_processamento = log.model_dump_json()
            importacao.data_fim_processamento = datetime.now()
            
            # Definir status final
            if len(log.erros) == 0 and len(log.sucessos) > 0:
                importacao.status = StatusImportacao.SUCESSO
            elif len(log.sucessos) > 0 and len(log.erros) > 0:
                importacao.status = StatusImportacao.SUCESSO_PARCIAL
            else:
                importacao.status = StatusImportacao.ERRO
                importacao.mensagem_erro = "Nenhum insumo foi importado com sucesso"
            
            self.db.commit()
            
            return True, f"Importação concluída: {len(log.sucessos)} sucessos, {len(log.erros)} erros"
            
        except Exception as e:
            logger.error(f"Erro ao processar importação: {e}")
            importacao.status = StatusImportacao.ERRO
            importacao.mensagem_erro = str(e)
            importacao.data_fim_processamento = datetime.now()
            self.db.commit()
            return False, f"Erro ao processar: {str(e)}"